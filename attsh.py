# version:  1.0
# date:     2023-05-24
# author:   Lijiakuan
# contact:  lijiakuan1988@gmail.com
# pip install pipreqs  pipreqs ./
# pip install streamlit
# pip install pandas
# pip install sqlalchemy
# pip install pymysql
# pip install Pillow
# pip install python-archive
import os
import sqlite3
import streamlit as st
import pandas as pd
from PIL import Image
from io import BytesIO
import base64
import time
# import PyPDF2

# import shutil
import datetime
# from pathlib import Path
import zipfile
from PyPDF2 import PdfReader

from openpyxl.utils.dataframe import dataframe_to_rows
import docx2txt
import plotly.express as px
# import plotly.graph_objects as go
# from plotly.subplots import make_subplots


#前端  
#整体配置

st.set_page_config(
    page_title="五中心医工科文件信息系统",
    page_icon='🌐',
    layout="wide",
    menu_items={
    'Get Help': 'https://github.com/',
    'About': '关于本系统: **由李家宽制作**'
}
)
UPLOAD_FOLDER = "uploads"
# 创建数据库连接
def create_connection():
    conn = sqlite3.connect("file_database.db")
    return conn

#创建数据库
@st.cache_resource
def create_table():
    conn = create_connection()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS files
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_title TEXT,
                file_pages INTEGER,
                creator TEXT,
                contact TEXT,
                hgroup TEXT,
                file_reciper TEXT,
                file_saver TEXT,
                entry_time TIMESTAMP,
                file_summary TEXT,
                file_name TEXT,
                file_path TEXT,
                remarks TEXT)''')
    conn.commit()
    conn.close()

create_table()
# 附件上传下载

def save_uploaded_file(uploaded_file,path):
    with open(os.path.join(path, uploaded_file.name), "wb") as f:
        f.write(uploaded_file.getbuffer())

def handle_file_upload():
    uploaded_files = st.file_uploader("附件", accept_multiple_files=True)
    file_details = []
    for uploaded_file in uploaded_files:
        save_uploaded_file(uploaded_file,UPLOAD_FOLDER)
        file_detail = {
            "FileName": uploaded_file.name,
            "FileType": uploaded_file.type,
            "FileSize": uploaded_file.size,
            "file_path": os.path.join(UPLOAD_FOLDER, uploaded_file.name)
        }
        file_details.append(file_detail)
    for file in file_details:
        st.write("成功上传文件:", file['FileName'])
    return file_details
    
        
#记录增删改查
def insert_record(record):
    conn = create_connection()
    c = conn.cursor()
    c.execute("INSERT INTO files (file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks) VALUES (?, ?, ?,?, ?, ?, ?, ?, ?, ?, ?, ?)",
              (record['file_title'], record['file_pages'],record['creator'] , record['contact'], record['hgroup'], 
               record['file_reciper'], record['file_saver'], record['entry_time'], record['file_summary'], record['file_name'], record['file_path'], record['remarks'],))
    conn.commit()
    conn.close()

def update_record(id, file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks):
    conn = create_connection()
    c = conn.cursor()
    c.execute("UPDATE files SET file_title=?, file_pages=?, creator=?, contact=?, hgroup=?, file_reciper=?, file_saver=?, entry_time=?, file_summary=?,file_name=?, file_path=?, remarks=? WHERE id=?",
              (file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks, id))
    conn.commit()
    conn.close()

def delete_record(id):
    # st.warning("Warning")
    conn = create_connection()
    c = conn.cursor()
    c.execute("DELETE FROM files WHERE id=?", (id,))
    conn.commit()
    conn.close()

def get_all_records():
    conn = create_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM files")
    records = c.fetchall()
    conn.close()
    return records

def get_record_by_id(id):
    conn = create_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM files WHERE id=?", (id,))
    record = c.fetchone()
    conn.close()
    return record

def get_record_by_ids(ids):
    conn = create_connection()
    c = conn.cursor()
    wstr = ("?, " * len(ids))[:-2]
    # print("this is wstr",wstr)
    query = f"SELECT * FROM files WHERE id IN  ({wstr})"
    params = tuple(ids)
    # print("this is params:",params)
    # d.execute(query,ids)
    c.execute(query, ([str(x) for x in ids]))
    serecords1 = c.fetchall()
    # print("this is serecords1",serecords1)
    conn.commit()
    conn.close()
    return serecords1
#关键词检索功能
def search_records(keyword):
    conn = create_connection()
    c = conn.cursor()
    # file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks
    c.execute("SELECT * FROM files WHERE file_name LIKE ? OR file_title LIKE ? OR creator LIKE ?  OR contact LIKE ? OR remarks LIKE ? OR file_summary LIKE ? OR hgroup LIKE ? OR file_reciper LIKE ? OR file_saver LIKE ?",
              (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
    records = c.fetchall()
    conn.close()
    return records
#关键词检索功能加日期限定
def search_records_with_date(keyword,start_date,end_date):
    conn = create_connection()
    c = conn.cursor()
    if keyword:
        # c.execute(f"SELECT * FROM files WHERE file_name LIKE ? OR file_title LIKE ? OR creator LIKE ?  OR contact LIKE ? OR remarks LIKE ? OR file_summary LIKE ? AND strftime('%Y-%m-%d',entry_time) between  strftime('%Y-%m-%d',{start_date}) AND  strftime('%Y-%m-%d',{end_date})",
        #         (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
        c.execute("SELECT * FROM files WHERE (file_name LIKE ? OR file_title LIKE ? OR creator LIKE ? OR contact LIKE ? OR remarks LIKE ? OR file_summary LIKE ? OR hgroup LIKE ? OR file_reciper LIKE ? OR file_saver LIKE ?) AND strftime('%Y-%m-%d', entry_time) BETWEEN ? AND ?",
            (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
    else:
        # c.execute(f"SELECT * FROM files WHERE strftime('%Y-%m-%d',entry_time) between strftime('%Y-%m-%d',{start_date}) AND strftime('%Y-%m-%d',{end_date})")
        c.execute("SELECT * FROM files WHERE  strftime('%Y-%m-%d', entry_time) BETWEEN ? AND ?",
            (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
    records = c.fetchall()
    conn.close()
    return records
#预览附件
def display_file(file_path):
    file_type = file_path.split(".")[-1].lower()
    # print (file_type)

    if file_type in ["jpg", "jpeg", "png","PNG"]:
        image = Image.open(file_path)
        st.image(image, caption=file_path, use_column_width=True)
    elif file_type in ["zip"]:
        with zipfile.ZipFile(file_path, "r") as z:
            st.write("Zip Content:")
            for f in z.namelist():
                st.write(f)
    elif file_type in ["pdf", "PDF"]:
        try:
            pdf = PdfReader(file_path, "rb")
            page = pdf.pages[1] 
            text = page.extract_text()
            st.write(text)
        except:
            st.warning("预览的PDF文档不能为扫描件！！")
        

    elif file_type in ["txt", "TXT","csv","CSV"]:
        with open(file_path, "rb") as f:
            text = str(f.read(),"utf-8")
            st.write(text)
            f.close()
    elif file_type in ["doc", "docx"]:
        docx_text = docx2txt.process(file_path)
        st.write(docx_text)
                    
    else:
        st.write("Unsupported file format.")

#记录导出功能
def export_records(records):
    datestr = datetime.datetime.now()
    date_str = datestr.strftime('%Y_%m_%d')
    # expfilpth = f"downloads/记录{date_str}.xlsx"

    df = pd.DataFrame(records, columns=["id", "文件标题","文件总页数","文件上报人","上报人联系方式","所属组", "文件接收人",  "文件保管人", "审批完成日期", "文件摘要", "文件名", "文件存储路径","备注"])
    df.to_excel(f"downloads/记录{date_str}.xlsx", index=False)
    # return expfilpth
#导出选中的记录功能
def export_selected_records(records):
    serecords = get_record_by_ids(records)
    # print(serecords) 
    datestr = datetime.datetime.now()
    date_str = datestr.strftime('%Y_%m_%d')
    # expfilpth = f"downloads/记录{date_str}.xlsx"
    df = pd.DataFrame(serecords, columns=["id", "文件标题","文件总页数","文件上报人","上报人联系方式","所属组", "文件接收人",  "文件保管人", "审批完成日期", "文件摘要", "文件名", "文件存储路径","备注"])
    df.to_excel(f"downloads/记录{date_str}.xlsx", index=False)    
    
# 导出Excel按钮
# def export_excel():
#     wb = Workbook()
#     ws = wb.active
#     date_str = end_time.get()
#     date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')
#     new_date_str = date_obj.strftime('%Y_%m_%d')
#     ws.title = '血压数据'

#     # 写入表头
#     for col in cols:
#         ws.cell(row=1, column=cols.index(col)+1).value = col

#     # 写入数据
#     for row in result:
#         for i in range(len(row)):
#             ws.cell(row=result.index(row)+2, column=i+1).value = str(row[i])

#     wb.save('血压数据{}.xlsx'.format(new_date_str))
#     messagebox.showinfo('提示', '导出Excel成功!')
    

def get_image_download_link(img, filename):
    buffered = BytesIO()
    img.save(buffered, format="JPEG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    href = f'<a href="data:file/jpeg;base64,{img_str}" download="{filename}" target="_blank">点击下载图片</a>'
    return href
        

  

st.title("文件信息系统系统")
#侧边栏
st.sidebar.header("医工科文件信息系统")
mode = st.sidebar.selectbox("附件管理", ["新增记录", "查看记录", "检索记录","图表统计展示"])
st.image("./banner1.png",use_column_width='always')
# 上传文件
############################################# 第一页 ############################################
if mode == "新增记录":
    st.header("🏢新增记录✍")
    filedetails = handle_file_upload()
    if filedetails:
        filenames = [d['FileName'] for d in filedetails]
        filepaths = [d['file_path'] for d in filedetails]

        st.header("添加记录")

        file_title = st.text_input("文件标题")
        page_count = st.number_input("文件总页数", min_value=1, value=1, step=1)
        creator = st.text_input("文件上报人")
        contact = st.text_input("上报人联系方式")
        hgroup = st.selectbox("所属组别",["维修组", "计量组", "库房组", "管理组", "采购组", "其它"])
        file_reciper = st.text_input("文件接收人")
        file_saver = st.text_input("文件保管人")
        entry_time = st.date_input("审批完成时间")
        summary = st.text_area("文件摘要")
        
        remarks = st.text_area("备注信息")
        

        if st.button("保存记录"):
            record = {
                "file_title": file_title,
                "file_pages": page_count,
                "creator": creator,
                "contact": contact,
                "hgroup":hgroup,
                "file_reciper":file_reciper,
                "file_saver":file_saver,
                "entry_time": entry_time,
                "file_summary": summary,
                "file_name": filenames[0],
                "file_path": filepaths[0],
                "remarks": remarks         
            }
            if record["file_title"] and record["contact"] is not None:
                insert_record(record)
                st.success("记录已保存")
                st.empty()
            else:
                st.warning("文件标题和联系方式不能为空")

######################################### 第二页 ##########################################
# 显示记录
if mode == "查看记录":
    PAGE_SIZE = 6
    st.header("👀查看记录👣")
    records = get_all_records()    
    if records:
        # st.write(df)
        current_page = st.number_input("当前页数 (从0开始)", min_value=0, max_value=(len(records) // PAGE_SIZE ) ,step=1)
        start_index = current_page * PAGE_SIZE
        end_index = start_index + PAGE_SIZE
        selected_records = []
        df = pd.DataFrame(records,columns=["id", "文件标题","文件总页数","文件上报人","上报人联系方式","所属组", "文件接收人", "文件保管人", "审批完成日期", "文件摘要", "文件名", "文件存储路径","备注"])
        # st.write(df.to_html(doctype=dt.HTMLFormatter(classes='display compact')))
        # df = pd.DataFrame(columns=["id", "file_name", "file_title", "entry_time", "creator", "contact", "file_summary", "file_pages", "remarks", "file_path"])
        df = df.reset_index(drop=True)
        df["复选框"] = ""
        col1, col2 = st.columns([0.1,1])
        col1.markdown(f'<div style="margin-top: 2.6rem;line-height:50px"></div>', unsafe_allow_html=True)
        col2.markdown(f'<div style="line-height:30px"></div>', unsafe_allow_html=True)
        # col1.markdown('<style>div.stCheckbox> {line-height:0rem;}</style>', unsafe_allow_html=True)
        if len(records) > PAGE_SIZE:
            total_pages = len(records) // PAGE_SIZE + 1
            st.write(f"共 {len(records)} 条记录，当前显示第 {current_page+1} 页 / 共 {total_pages} 页")
        if ( len(records)-start_index >0 and len(records)-start_index< PAGE_SIZE):
            for i in range(start_index, len(records)):
                df.at[i, "复选框"] = col1.checkbox(f"{i}",key=f"record_{df.at[i, 'id']}")
                if df.at[i, "复选框"]:
                    selected_records.append(df.at[i, 'id'])
            col2.write(df[start_index:len(records)])
        elif (len(records)-start_index > PAGE_SIZE):
            for i in range(start_index, end_index):
                df.at[i, "复选框"] = col1.checkbox(f"{i}",key=f"record_{df.at[i, 'id']}")
                if df.at[i, "复选框"]:
                    selected_records.append(df.at[i, 'id'])
            col2.write(df[start_index:end_index])
    else:
        st.write("没有记录")
    # print(selected_records)
    #导出所有记录
    if st.button("导出全部记录",on_click=export_records(records)):
        datestr = datetime.datetime.now()
        date_str = datestr.strftime('%Y_%m_%d')
        expfilpth = f"downloads\记录{date_str}.xlsx"
        exfilname = f"记录{date_str}.xlsx"
        exp_btn = st.download_button(
                        label="下载记录",
                        data=open(expfilpth, "rb"),
                        file_name=exfilname,
                        )
    #导出选中的记录
    
    if st.button("导出选中记录"):
        if selected_records:
            export_selected_records(selected_records)
            datestr = datetime.datetime.now()
            date_str = datestr.strftime('%Y_%m_%d')
            expfilpth = f"downloads/记录{date_str}.xlsx"
            exfilname = f"记录{date_str}.xlsx"
            exp_btn = st.download_button(
                            label="下载记录",
                            data=open(expfilpth, "rb"),
                            file_name=exfilname,
                            )       
    
        
    cl1, cl2, cl3,cl4 = st.columns([0.2,0.2,0.2,1])
    file_path = ''
    #预览文件
    with cl1:
        if st.button("预览文件"):
            if len(selected_records) == 1:
                idr1 = int(selected_records[0])
                filename = get_record_by_id(idr1)[-2]
                file_path = os.path.join(os.getcwd(),filename)
                # st.write(file_path)
            else:
                st.warning("请选择一条记录！") 

        
    if file_path:
        display_file(str(file_path))
        
    #下载附件
    with cl2:
        if len(selected_records) == 1:
            idr2 = int(selected_records[0])
            filpth = get_record_by_id(idr2)[-2]
            filpth1 = r'{}'.format(filpth)
            filna = filpth1.split('\/')[1]
            down_btn = st.download_button(
                    label="下载",
                    data=open(filpth, "rb"),
                    file_name=filna
                    )
        
    #编辑记录
    
    with cl4:
        # if st.button("编辑记录"):
        if len(selected_records) == 1:                 
            idr2 = int(selected_records[0])
            row_selected = get_record_by_id(idr2)
            with st.container():
                my_form1 = st.form(key='my_form')
                my_form1.header("修改记录")
                # st.write(row[2])
                
                uploaded_files = my_form1.file_uploader("请上传附件", accept_multiple_files=True)
                title = my_form1.text_input("文件标题", value=row_selected[1])
                page_count = my_form1.number_input(
                    "文件页数", min_value=1, value=row_selected[2], step=1)
                creator = my_form1.text_input("文件上报人", value=row_selected[3])
                contact = my_form1.text_input("上报人联系方式", value=row_selected[4])
                hgroup = my_form1.text_input("所属组别", value=row_selected[5])
                file_reciper = my_form1.text_input("文件接收人", value=row_selected[6])
                file_saver = my_form1.text_input("文件保管人", value=row_selected[7])
                entry_time2 = my_form1.text_input("审批完成时间",value=row_selected[8])
                try:
                    entry_time2 = datetime.datetime.strptime(entry_time2, "%Y-%m-%d") 
                except ValueError:
                    entry_time2 =  entry_time2[:-9]      
                    entry_time2 = datetime.datetime.strptime(entry_time2, "%Y-%m-%d")                       
                summary = my_form1.text_input("文件摘要", value=row_selected[9])

                remarks = my_form1.text_input("备注信息", value=row_selected[12])
                
                if uploaded_files:
                    for uploaded_file in uploaded_files:
                        save_uploaded_file(uploaded_file,UPLOAD_FOLDER)
                    fil1 = uploaded_files[0]
                    fname = st.write(f"文件名:{fil1.name}")
                    fpath = os.path.join(UPLOAD_FOLDER, fil1.name)
                else:
                    fname = my_form1.write(f"文件名:{row_selected[10]}")
                    fpath = row_selected[11]
                submit_button = my_form1.form_submit_button("提交修改")
                if  submit_button:
                    # print(submit_button)
                    new_record = {
                        "title":title,
                        "page_count":page_count,
                        "creator":creator,
                        "contact": contact,
                        "hgroup":hgroup,
                        "file_reciper":file_reciper,
                        "file_saver":file_saver,
                        "entry_time": entry_time2,
                        "summary": summary,
                        "file_name": fname,
                        "file_path": fpath,
                        "remarks": remarks                        
                    }
                    update_record(idr2, new_record["title"],new_record["page_count"], new_record["creator"], new_record["contact"], new_record["hgroup"]
                                  , new_record["file_reciper"], new_record["file_saver"], new_record["entry_time"], new_record["summary"], new_record["file_name"], new_record["file_path"], new_record["remarks"])
                    # print(idr2, new_record["name"],new_record["title"], new_record["entry_time"], new_record["creator"], new_record["contact"]
                    #                 , new_record["summary"], new_record["page_count"], new_record["remarks"], new_record["filepath"])
                    st.success("记录已更新") 
                    time.sleep(2)
                    st.experimental_rerun()
        elif len(selected_records) > 1: 
            st.warning("🚨请选择一条记录进行修改🚨")
                
    #删除记录
    with cl3:
        if st.button("删除记录"):
            if len(selected_records) >= 1 :
                for record in selected_records:
                    idr3 = int(record)
                    delete_record(idr3)
                st.success("删除记录成功！")
                st.balloons()
                time.sleep(3)
                st.experimental_rerun()  

######################################### 第三页 ##########################################
if mode == "检索记录":
    PAGE_SIZE = 6
    
    # df = pd.DataFrame(records,columns=["id", "文件名", "文件标题", "日期", "创建人", "创建人联系方式", "文件摘要", "页码数", "备注", "附件存放地址"])
    # my_form = st.form(key='search_keywords')    
    # st.header("检索记录")
    # search_type = st.selectbox("检索类型", ["内容检索", "时间区间"])
    # keyword = st.text_input("关键词")
    # if keyword:
    #     records = search_records(keyword)
    #     run_search(records)
    
    # if search_type == "时间区间":
    #     start_date = st.date_input("开始日期")
    #     end_date = st.date_input("结束日期")
    # search_records(keyword)
    # if my_form.form_submit_button("检索"):
    #     records = search_records(keyword)
    # else:
    #     records = ""
    # if records:
        # st.write(df)
    def run_search(records):
        records = records
        current_page = st.number_input("当前页数 (从0开始)", min_value=0, max_value=(len(records) // PAGE_SIZE ) ,step=1)
        start_index = current_page * PAGE_SIZE
        end_index = start_index + PAGE_SIZE
        selected_records = []
        #file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks
        df = pd.DataFrame(records,columns=["id","文件标题","文件总页数","文件上报人","上报人联系方式","所属组", "文件接收人", "文件保管人", "审批完成日期", "文件摘要", "文件名", "文件存储路径","备注"])
        # st.write(df.to_html(doctype=dt.HTMLFormatter(classes='display compact')))
        # df = pd.DataFrame(columns=["id", "file_name", "file_title", "entry_time", "creator", "contact", "file_summary", "file_pages", "remarks", "file_path"])
        df = df.reset_index(drop=True)
        df["复选框"] = ""
        col1, col2 = st.columns([0.1,1])
        col1.markdown(f'<div style="margin-top: 2.6rem;line-height:50px"></div>', unsafe_allow_html=True)
        col2.markdown(f'<div style="line-height:30px"></div>', unsafe_allow_html=True)
        # col1.markdown('<style>div.stCheckbox> {line-height:0rem;}</style>', unsafe_allow_html=True)
        if len(records) > PAGE_SIZE:
            total_pages = len(records) // PAGE_SIZE + 1
            st.write(f"共 {len(records)} 条记录，当前显示第 {current_page+1} 页 / 共 {total_pages} 页")
        if ( len(records)-start_index >0 and len(records)-start_index< PAGE_SIZE):
            for i in range(start_index, len(records)):
                df.at[i, "复选框"] = col1.checkbox(f"{i}",key=f"record_{df.at[i, 'id']}")
                if df.at[i, "复选框"]:
                    selected_records.append(df.at[i, 'id'])
            col2.write(df[start_index:len(records)])
        elif (len(records)-start_index > PAGE_SIZE):
            for i in range(start_index, end_index):
                df.at[i, "复选框"] = col1.checkbox(f"{i}",key=f"record_{df.at[i, 'id']}")
                if df.at[i, "复选框"]:
                    selected_records.append(df.at[i, 'id'])
            col2.write(df[start_index:end_index])
        # else:
        #     st.write("没有记录")
        # print(selected_records)
        #导出所有记录
        if st.button("导出全部记录",on_click=export_records(records)):
            datestr = datetime.datetime.now()
            date_str = datestr.strftime('%Y_%m_%d')
            expfilpth = f"downloads/记录{date_str}.xlsx"
            exfilname = f"记录{date_str}.xlsx"
            exp_btn = st.download_button(
                            label="下载记录",
                            data=open(expfilpth, "rb"),
                            file_name=exfilname,
                            )
        #导出选中的记录
        
        if st.button("导出选中记录"):
            if selected_records:
                export_selected_records(selected_records)
                datestr = datetime.datetime.now()
                date_str = datestr.strftime('%Y_%m_%d')
                expfilpth = f"downloads/记录{date_str}.xlsx"
                exfilname = f"记录{date_str}.xlsx"
                exp_btn = st.download_button(
                                label="下载记录",
                                data=open(expfilpth, "rb"),
                                file_name=exfilname,
                                )       
        
            
        cl1, cl2, cl3,cl4 = st.columns([0.2,0.2,0.2,1])
        file_path = ''
        #预览文件
        with cl1:
            if st.button("预览文件"):
                if len(selected_records) == 1:
                    idr1 = int(selected_records[0])
                    filename = get_record_by_id(idr1)[-2]
                    file_path = os.path.join(os.getcwd(),filename)
                    # st.write(file_path)
                else:
                    st.warning("请选择一条记录！") 

            
        if file_path:
            display_file(str(file_path))
            
        #下载附件
        with cl2:
            if len(selected_records) == 1:
                idr2 = int(selected_records[0])
                filpth = get_record_by_id(idr2)[-2]
                filpth1 = r'{}'.format(filpth) 
                filna = filpth1.split('\/')[1]
                # print()
                down_btn = st.download_button(
                        label="下载",
                        data=open(filpth, "rb"),
                        file_name=filna
                        )
            
        #编辑记录
        
        with cl4:
            # if st.button("编辑记录"):
            if len(selected_records) == 1:                 
                idr2 = int(selected_records[0])
                row_selected = get_record_by_id(idr2)
                with st.container():
                    my_form1 = st.form(key='my_form')
                    my_form1.header("修改记录")
                    # st.write(row[2])
                    
                    uploaded_files = my_form1.file_uploader("请上传附件", accept_multiple_files=True)
                    title = my_form1.text_input("文件标题", value=row_selected[1])
                    page_count = my_form1.number_input("文件页数", min_value=1, value=row_selected[2], step=1)
                    creator = my_form1.text_input("文件上报人", value=row_selected[3])
                    contact = my_form1.text_input("上报人联系方式", value=row_selected[4])
                    hgroup = my_form1.text_input("所属组别", value=row_selected[5])
                    file_reciper = my_form1.text_input("文件接收人", value=row_selected[6])
                    file_saver = my_form1.text_input("文件保管人", value=row_selected[7])
                    # entry_time = my_form1.date_input(value=row_selected[8])
                    entry_time1 = my_form1.text_input("审批完成时间",value=row_selected[8])
                    try:
                        entry_time1 = datetime.datetime.strptime(entry_time1, "%Y-%m-%d") 
                    except ValueError:
                        entry_time1 =  entry_time1[:-9]      
                        entry_time1 = datetime.datetime.strptime(entry_time1, "%Y-%m-%d")    
                    summary = my_form1.text_input("文件摘要", value=row_selected[9])

                    remarks = my_form1.text_input("备注信息", value=row_selected[12])
                    
                    if uploaded_files:
                        for uploaded_file in uploaded_files:
                            save_uploaded_file(uploaded_file,UPLOAD_FOLDER)
                        fil1 = uploaded_files[0]
                        fname = st.write(f'**文件名**{fil1.name}')
                        fpath = os.path.join(UPLOAD_FOLDER, fil1.name)
                    else:
                        fname = my_form1.write(f'**文件名**{row_selected[10]}' )
                        fpath = row_selected[11]
                    submit_button = my_form1.form_submit_button("提交修改")
                    if  submit_button:
                        # print(submit_button)
                        new_record = {
                            "title":title,
                            "page_count":page_count,
                            "creator":creator,
                            "contact": contact,
                            "hgroup":hgroup,
                            "file_reciper":file_reciper,
                            "file_saver":file_saver,
                            "entry_time": entry_time1,
                            "summary": summary,
                            "file_name": fname,
                            "file_path": fpath,
                            "remarks": remarks                        
                        }
                        update_record(idr2, new_record["title"],new_record["page_count"], new_record["creator"], new_record["contact"], new_record["hgroup"]
                                    , new_record["file_reciper"], new_record["file_saver"], new_record["entry_time"], new_record["summary"], new_record["file_name"], new_record["file_path"], new_record["remarks"])
                        # print(idr2, new_record["name"],new_record["title"], new_record["entry_time"], new_record["creator"], new_record["contact"]
                        #                 , new_record["summary"], new_record["page_count"], new_record["remarks"], new_record["filepath"])
                        st.success("记录已更新")
                        time.sleep(2)
                        st.experimental_rerun()
                         
            elif len(selected_records) > 1: 
                st.warning("🚨请选择一条记录进行修改🚨")
                    
        #删除记录
        with cl3:
            if st.button("删除记录"):
                if len(selected_records) >= 1 :
                    for record in selected_records:
                        idr3 = int(record)
                        delete_record(idr3)
                    st.success("删除记录成功！")
                    st.balloons()
                    time.sleep(2)
                    st.experimental_rerun()
    
    st.header("🔎检索记录🔬")
    search_type = st.selectbox("检索类型", ["内容检索", "时间区间"])
    keyword = st.text_input("关键词")
    if search_type == "内容检索":
        if keyword:
            records = search_records(keyword)
            run_search(records)
        else:
            st.warning("🚨请输入检索内容！🚨")
    
    if search_type == "时间区间":
        # my_form = st.form(key="filter_time")
        start_date = st.date_input("开始日期")
        end_date = st.date_input("结束日期")
        # submit_button = my_form.form_submit_button("查询")

        if (start_date and end_date or keyword):
            records1 = search_records_with_date(keyword,start_date,end_date)
            # print(keyword,start_date,end_date,records1)
            run_search(records1)

############################################# 第四页 ############################################
if mode == "图表统计展示":
    st.header("📊图表统计展示📈")
    records = get_all_records()
    #file_title, file_pages, creator, contact, hgroup, file_reciper, file_saver, entry_time, file_summary, file_name, file_path, remarks
    df = pd.DataFrame(records,columns=["id", "file_title", "file_pages", "creator", "contact", "组别", "file_reciper", "file_saver", "entry_time", "file_summary", "file_name", "file_path", "remarks"])
    df["entry_time"] = pd.to_datetime(df["entry_time"])  # 将 entry_time 列转换为 Pandas 的日期时间格式
    df["年份"] = df["entry_time"].apply(lambda x: x.year)  # 添加 year 列
    df["月份"] = df["entry_time"].apply(lambda x: x.month)  # 添加 month 列
    df["日期"] = df["entry_time"].dt.strftime("%Y-%m-%d")
    year_counts = df.groupby("年份").size().reset_index(name="年附件数")
    month_counts = df.groupby(["年份", "月份"]).size().reset_index(name="月附件数")
    div_counts = df.groupby(["年份", "月份", "组别"]).size().reset_index(name="月附件总数")
    div_re_counts = df.groupby(["年份", "组别"]).size().reset_index(name="年附件总数")
    date_div_counts = df.groupby(["日期", "组别"]).size().reset_index(name="日附件总数")
    # st.table(df)
    annue = st.selectbox("年度选择",(2021,2022,2023,2024,2025,2026),index=2)
    col11,col12 = st.columns(2)
    # 按年月统计图表       
    month_chart = px.pie(
        month_counts[month_counts["年份"]==annue],
        names = "月份",
        values = "月附件数",
        hole=0.25,
        title="附件记录月统计图"
        )
    col11.subheader("按月统计") 
    col11.plotly_chart(month_chart)
    #按年统计图表
    col12.subheader("按年统计") 
    year_chart = px.bar(year_counts, x="年份", y="年附件数", title="附件记录年统计图")
    col12.plotly_chart(year_chart)
    #按组别每月数量
    bar_graph = px.bar(
        div_counts[div_counts["年份"]== annue],
        x  = "月份",
        y = "月附件总数",
        title = "按组别统计每月附件数量",
        color="组别",
        # facet_col="remarks"
        )
    # 设置 x 轴和 y 轴标签
    bar_graph.update_xaxes(title_text="月份")
    bar_graph.update_yaxes(title_text="各组提报数量")
    col11.plotly_chart(bar_graph)
    
    #按组别数量
    
    bar_graph_zb = px.pie(
    div_re_counts[div_re_counts["年份"]==annue],
    names  = "组别",
    values = "年附件总数",
    title = "按组别统计每年附件数量",
    color="组别",
    # facet_col="remarks"
    )
    col12.plotly_chart(bar_graph_zb)
    
    line_chart = px.line(
    date_div_counts,
    x = "日期",
    y = "日附件总数",
    color = "组别"
    )
    st.header("Line Chart")
    st.plotly_chart(line_chart)
    
